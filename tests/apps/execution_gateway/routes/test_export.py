"""Tests for Excel export in apps/execution_gateway/routes/export.py.

Verifies that _generate_excel_content returns real grid data (not placeholders)
and that all cell values are sanitised against formula injection.
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any
from unittest.mock import MagicMock

import pytest

from apps.execution_gateway.app_factory import create_mock_context
from apps.execution_gateway.routes import export as export_module
from apps.execution_gateway.routes.export import (
    _GRID_COLUMNS,
    _build_filter_clause,
    _build_order_clause,
    _resolve_sort_aliases,
    _validate_columns,
)

# ---------------------------------------------------------------------------
# Unit tests for helper functions
# ---------------------------------------------------------------------------


class TestValidateColumns:
    def test_returns_allowlist_when_no_visible_columns(self) -> None:
        result = _validate_columns("positions", None)
        assert result == _GRID_COLUMNS["positions"]

    def test_filters_unknown_columns(self) -> None:
        result = _validate_columns("positions", ["symbol", "EVIL_COL", "qty"])
        assert result == ["symbol", "qty"]

    def test_preserves_requested_order(self) -> None:
        result = _validate_columns("orders", ["status", "symbol", "side"])
        assert result == ["status", "symbol", "side"]

    def test_falls_back_to_default_when_all_invalid(self) -> None:
        result = _validate_columns("fills", ["bad1", "bad2"])
        assert result == _GRID_COLUMNS["fills"]

    def test_resolves_frontend_column_aliases(self) -> None:
        """Frontend alias "time" is resolved to "executed_at" for fills."""
        result = _validate_columns("fills", ["time", "symbol", "qty"])
        assert result == ["executed_at", "symbol", "qty"]

    def test_resolves_orders_type_alias(self) -> None:
        """Frontend alias "type" is resolved to "order_type" for orders."""
        result = _validate_columns("orders", ["symbol", "type", "status"])
        assert result == ["symbol", "order_type", "status"]

    def test_tca_falls_back_when_mostly_computed_columns(self) -> None:
        """TCA grid sends computed columns (fill_rate_pct, is_bps etc.) alongside
        a few raw ones.  When fewer than half survive validation, fall back to
        the full allowlist so TCA exports are not misleadingly incomplete."""
        # Simulate TCA grid sending mostly computed columns (no aliases)
        result = _validate_columns(
            "tca",
            ["fill_rate_pct", "is_bps", "slippage_bps", "vwap_deviation", "symbol"],
        )
        # Only symbol is valid (1/5 < half), so full fallback
        assert result == _GRID_COLUMNS["tca"]

    def test_tca_execution_date_alias_resolved(self) -> None:
        """Frontend alias 'execution_date' is resolved to 'executed_at' for TCA."""
        result = _validate_columns(
            "tca",
            ["execution_date", "symbol", "side", "qty", "price"],
        )
        # All 5 resolve to valid columns (execution_date -> executed_at)
        assert result == ["executed_at", "symbol", "side", "qty", "price"]

    def test_no_fallback_when_majority_valid(self) -> None:
        """When the majority of requested columns are valid, keep the subset."""
        result = _validate_columns("orders", ["symbol", "side", "qty", "bad_col"])
        assert result == ["symbol", "side", "qty"]


class TestResolveSortAliases:
    def test_returns_none_when_no_sort_model(self) -> None:
        assert _resolve_sort_aliases("orders", None) is None

    def test_resolves_orders_type_alias(self) -> None:
        model = [{"colId": "type", "sort": "asc"}]
        result = _resolve_sort_aliases("orders", model)
        assert result is not None
        assert result[0]["colId"] == "order_type"

    def test_resolves_fills_time_alias(self) -> None:
        model = [{"colId": "time", "sort": "desc"}]
        result = _resolve_sort_aliases("fills", model)
        assert result is not None
        assert result[0]["colId"] == "executed_at"

    def test_no_change_for_grids_without_aliases(self) -> None:
        model = [{"colId": "symbol", "sort": "asc"}]
        result = _resolve_sort_aliases("positions", model)
        assert result is model  # Same object returned (no aliases to resolve)

    def test_preserves_sort_direction(self) -> None:
        model = [{"colId": "type", "sort": "desc"}]
        result = _resolve_sort_aliases("orders", model)
        assert result is not None
        assert result[0]["sort"] == "desc"

    def test_unknown_aliases_pass_through(self) -> None:
        model = [{"colId": "symbol", "sort": "asc"}, {"colId": "type", "sort": "desc"}]
        result = _resolve_sort_aliases("orders", model)
        assert result is not None
        assert result[0]["colId"] == "symbol"
        assert result[1]["colId"] == "order_type"


class TestBuildFilterClause:
    def test_returns_empty_when_no_filter(self) -> None:
        sql, params = _build_filter_clause(None, ["symbol"])
        assert sql == ""
        assert params == []

    def test_text_contains_filter(self) -> None:
        filt = {"symbol": {"filterType": "text", "type": "contains", "filter": "AAPL"}}
        sql, params = _build_filter_clause(filt, ["symbol"])
        assert "ILIKE" in sql
        assert "%AAPL%" in params

    def test_text_equals_filter(self) -> None:
        filt = {"status": {"filterType": "text", "type": "equals", "filter": "filled"}}
        sql, params = _build_filter_clause(filt, ["status"])
        assert "= %s" in sql
        assert "filled" in params

    def test_number_greater_than_filter(self) -> None:
        filt = {"qty": {"filterType": "number", "type": "greaterThan", "filter": 10}}
        sql, params = _build_filter_clause(filt, ["qty"])
        assert "> %s" in sql
        assert 10 in params

    def test_number_in_range_filter(self) -> None:
        filt = {"qty": {"filterType": "number", "type": "inRange", "filter": 5, "filterTo": 20}}
        sql, params = _build_filter_clause(filt, ["qty"])
        assert ">= %s" in sql
        assert "<= %s" in sql
        assert 5 in params
        assert 20 in params

    def test_date_in_range_filter(self) -> None:
        filt = {
            "created_at": {
                "filterType": "date",
                "type": "inRange",
                "dateFrom": "2026-01-01",
                "dateTo": "2026-02-01",
            }
        }
        result_sql, params = _build_filter_clause(filt, ["created_at"])
        # Range uses >= start and < end+1day for index-friendly comparisons
        assert ">= %s" in result_sql
        assert "interval '1 day'" in result_sql
        assert "2026-01-01" in params
        assert "2026-02-01" in params

    def test_ignores_unknown_columns(self) -> None:
        filt = {"evil_col": {"filterType": "text", "type": "contains", "filter": "x"}}
        sql, params = _build_filter_clause(filt, ["symbol"])
        assert sql == ""
        assert params == []

    def test_col_prefix_applied(self) -> None:
        filt = {"symbol": {"filterType": "text", "type": "equals", "filter": "AAPL"}}
        sql, params = _build_filter_clause(filt, ["symbol"], col_prefix="p.")
        assert "p.symbol" in sql
        assert "AAPL" in params

    def test_multiple_filters(self) -> None:
        filt = {
            "symbol": {"filterType": "text", "type": "equals", "filter": "AAPL"},
            "qty": {"filterType": "number", "type": "greaterThan", "filter": 5},
        }
        sql, params = _build_filter_clause(filt, ["symbol", "qty"])
        assert "AND" in sql
        assert len(params) == 2

    def test_non_dict_spec_skipped(self) -> None:
        filt = {"symbol": "not-a-dict"}
        sql, params = _build_filter_clause(filt, ["symbol"])
        assert sql == ""
        assert params == []

    def test_date_greater_than_or_equal_filter(self) -> None:
        filt = {
            "created_at": {
                "filterType": "date",
                "type": "greaterThanOrEqual",
                "dateFrom": "2026-03-01",
            }
        }
        result_sql, params = _build_filter_clause(filt, ["created_at"])
        assert ">= %s" in result_sql
        assert "2026-03-01" in params

    def test_date_less_than_or_equal_filter(self) -> None:
        filt = {
            "created_at": {
                "filterType": "date",
                "type": "lessThanOrEqual",
                "dateFrom": "2026-06-01",
            }
        }
        result_sql, params = _build_filter_clause(filt, ["created_at"])
        # lessThanOrEqual uses < date + 1 day to include the full day
        assert "interval '1 day'" in result_sql
        assert "2026-06-01" in params

    def test_col_prefix_map_overrides_col_prefix(self) -> None:
        """Per-column mapping takes precedence over blanket col_prefix."""
        filt = {
            "order_qty": {"filterType": "number", "type": "greaterThan", "filter": 5},
            "symbol": {"filterType": "text", "type": "equals", "filter": "AAPL"},
        }
        prefix_map = {"order_qty": "o.qty", "symbol": "t.symbol"}
        result_sql, params = _build_filter_clause(
            filt, ["order_qty", "symbol"], col_prefix="t.", col_prefix_map=prefix_map,
        )
        assert "o.qty" in result_sql
        assert "t.symbol" in result_sql
        assert 5 in params
        assert "AAPL" in params

    def test_col_prefix_map_fallback_to_prefix(self) -> None:
        """Columns not in col_prefix_map fall back to col_prefix."""
        filt = {"symbol": {"filterType": "text", "type": "equals", "filter": "AAPL"}}
        prefix_map = {"order_qty": "o.qty"}  # symbol not mapped
        result_sql, params = _build_filter_clause(
            filt, ["symbol"], col_prefix="t.", col_prefix_map=prefix_map,
        )
        assert "t.symbol" in result_sql
        assert "AAPL" in params


class TestBuildOrderClause:
    def test_default_order_when_no_sort_model(self) -> None:
        assert _build_order_clause(None, ["symbol"], "symbol ASC") == "symbol ASC"

    def test_valid_sort_model(self) -> None:
        model = [{"colId": "symbol", "sort": "desc"}]
        result = _build_order_clause(model, ["symbol", "qty"], "symbol ASC")
        assert result == "symbol DESC"

    def test_ignores_unknown_columns(self) -> None:
        model = [{"colId": "DROP TABLE", "sort": "asc"}]
        result = _build_order_clause(model, ["symbol"], "symbol ASC")
        assert result == "symbol ASC"

    def test_sort_index_determines_precedence(self) -> None:
        """Multi-column sorts should respect sortIndex for precedence."""
        model = [
            {"colId": "qty", "sort": "asc", "sortIndex": 1},
            {"colId": "symbol", "sort": "desc", "sortIndex": 0},
        ]
        result = _build_order_clause(model, ["symbol", "qty"], "symbol ASC")
        assert result == "symbol DESC, qty ASC"

    def test_sort_index_missing_appended_at_end(self) -> None:
        """Items without sortIndex are appended after indexed items."""
        model = [
            {"colId": "qty", "sort": "asc"},
            {"colId": "symbol", "sort": "desc", "sortIndex": 0},
        ]
        result = _build_order_clause(model, ["symbol", "qty"], "symbol ASC")
        assert result == "symbol DESC, qty ASC"


# ---------------------------------------------------------------------------
# Integration tests for _generate_excel_content
# ---------------------------------------------------------------------------


def _make_cursor_mock(rows: list[tuple[Any, ...]], col_names: list[str]) -> MagicMock:
    """Create a mock cursor that returns the given rows."""
    cursor = MagicMock()
    cursor.description = [(name,) for name in col_names]
    cursor.fetchall.return_value = rows
    cursor.__enter__ = MagicMock(return_value=cursor)
    cursor.__exit__ = MagicMock(return_value=False)
    return cursor


def _make_conn_mock(cursor: MagicMock) -> MagicMock:
    conn = MagicMock()
    conn.cursor.return_value = cursor
    conn.__enter__ = MagicMock(return_value=conn)
    conn.__exit__ = MagicMock(return_value=False)
    return conn


def _make_ctx_with_rows(
    rows: list[tuple[Any, ...]], col_names: list[str]
) -> MagicMock:
    cursor = _make_cursor_mock(rows, col_names)
    conn = _make_conn_mock(cursor)
    ctx = create_mock_context()
    ctx.db.transaction.return_value = conn
    return ctx


@pytest.mark.asyncio()
class TestGenerateExcelContent:
    async def test_positions_returns_real_data(self) -> None:
        rows = [
            ("AAPL", Decimal("10"), Decimal("150.25"), Decimal("152.00"), Decimal("17.50"), Decimal("0"), datetime(2026, 1, 1, tzinfo=UTC)),
            ("MSFT", Decimal("5"), Decimal("400.00"), Decimal("405.00"), Decimal("25.00"), Decimal("10"), datetime(2026, 1, 2, tzinfo=UTC)),
        ]
        col_names = ["symbol", "qty", "avg_entry_price", "current_price", "unrealized_pl", "realized_pl", "updated_at"]
        ctx = _make_ctx_with_rows(rows, col_names)

        content, row_count = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="positions",
            strategy_ids=["alpha"],
            filter_params=None,
            visible_columns=None,
            sort_model=None,
        )

        assert row_count == 2
        assert isinstance(content, bytes)
        assert len(content) > 0

        # Verify it's a valid Excel file by reading it
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        assert ws.title == "Positions"
        # Header row
        assert ws.cell(1, 1).value == "symbol"
        # Data — NOT placeholder text
        assert ws.cell(2, 1).value == "AAPL"
        assert ws.cell(3, 1).value == "MSFT"
        assert "placeholder" not in str(ws.cell(2, 1).value).lower()
        assert "pending" not in str(ws.cell(2, 1).value).lower()

    async def test_orders_returns_real_data(self) -> None:
        rows = [
            ("ord-1", "alpha", "AAPL", "buy", 10, "market", None, None, "day", "filled", Decimal("10"), Decimal("150"), datetime(2026, 1, 1, tzinfo=UTC), datetime(2026, 1, 1, 0, 1, tzinfo=UTC)),
        ]
        col_names = ["client_order_id", "strategy_id", "symbol", "side", "qty", "order_type", "limit_price", "stop_price", "time_in_force", "status", "filled_qty", "filled_avg_price", "created_at", "filled_at"]
        ctx = _make_ctx_with_rows(rows, col_names)

        content, row_count = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="orders",
            strategy_ids=["alpha"],
            filter_params=None,
            visible_columns=["symbol", "side", "qty", "status"],
            sort_model=None,
        )

        assert row_count == 1
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        # Only requested (valid) columns appear
        assert ws.cell(1, 1).value == "symbol"
        assert ws.cell(1, 2).value == "side"

    async def test_unsupported_grid_raises(self) -> None:
        ctx = create_mock_context()
        with pytest.raises(NotImplementedError, match="not implemented"):
            await export_module._generate_excel_content(
                ctx=ctx,
                grid_name="unknown_grid",
                strategy_ids=["alpha"],
                filter_params=None,
                visible_columns=None,
                sort_model=None,
            )

    async def test_empty_result_returns_zero_row_count(self) -> None:
        ctx = _make_ctx_with_rows([], ["symbol", "qty", "avg_entry_price", "current_price", "unrealized_pl", "realized_pl", "updated_at"])

        content, row_count = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="positions",
            strategy_ids=["alpha"],
            filter_params=None,
            visible_columns=None,
            sort_model=None,
        )

        assert row_count == 0

    async def test_formula_injection_sanitised(self) -> None:
        """Values starting with = are prefixed with ' to prevent formula injection."""
        rows = [("=IMPORTDATA(url)", Decimal("1"), Decimal("1"), Decimal("1"), Decimal("0"), Decimal("0"), datetime(2026, 1, 1, tzinfo=UTC))]
        col_names = ["symbol", "qty", "avg_entry_price", "current_price", "unrealized_pl", "realized_pl", "updated_at"]
        ctx = _make_ctx_with_rows(rows, col_names)

        content, _ = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="positions",
            strategy_ids=["alpha"],
            filter_params=None,
            visible_columns=None,
            sort_model=None,
        )

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        # The dangerous formula should be prefixed with '
        assert ws.cell(2, 1).value == "'=IMPORTDATA(url)"

    async def test_numeric_types_preserved_in_excel(self) -> None:
        """Numeric and datetime values are stored as native types, not strings."""
        rows = [
            (
                "AAPL",
                Decimal("10"),
                Decimal("150.25"),
                Decimal("152.00"),
                Decimal("17.50"),
                Decimal("0"),
                datetime(2026, 1, 1, tzinfo=UTC),
            ),
        ]
        col_names = [
            "symbol", "qty", "avg_entry_price", "current_price",
            "unrealized_pl", "realized_pl", "updated_at",
        ]
        ctx = _make_ctx_with_rows(rows, col_names)

        content, _ = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="positions",
            strategy_ids=["alpha"],
            filter_params=None,
            visible_columns=None,
            sort_model=None,
        )

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        # String column stays string
        assert ws.cell(2, 1).value == "AAPL"
        # Numeric columns must NOT be stringified — openpyxl stores Decimal
        # as a number, so the cell value should be numeric (not a string).
        qty_val = ws.cell(2, 2).value
        assert not isinstance(qty_val, str), f"qty should be numeric, got str: {qty_val!r}"
        # datetime should be preserved as a datetime
        updated_val = ws.cell(2, 7).value
        assert not isinstance(updated_val, str), (
            f"updated_at should be datetime, got str: {updated_val!r}"
        )

    async def test_date_objects_preserved_in_excel(self) -> None:
        """PostgreSQL DATE columns (datetime.date) should be stored natively."""
        rows = [
            (
                "AAPL",
                Decimal("10"),
                Decimal("150.25"),
                Decimal("152.00"),
                Decimal("17.50"),
                Decimal("0"),
                date(2026, 1, 1),
            ),
        ]
        col_names = [
            "symbol", "qty", "avg_entry_price", "current_price",
            "unrealized_pl", "realized_pl", "updated_at",
        ]
        ctx = _make_ctx_with_rows(rows, col_names)

        content, _ = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="positions",
            strategy_ids=["alpha"],
            filter_params=None,
            visible_columns=None,
            sort_model=None,
        )

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        # date object should be preserved as a date, not stringified
        updated_val = ws.cell(2, 7).value
        assert not isinstance(updated_val, str), (
            f"updated_at should be date, got str: {updated_val!r}"
        )

    async def test_filter_params_applied_to_query(self) -> None:
        """Verify that filter_params produce SQL WHERE clauses in the query."""
        rows = [
            ("ord-1", "alpha", "AAPL", "buy", 10, "market", None, None, "day", "filled", Decimal("10"), Decimal("150"), datetime(2026, 1, 1, tzinfo=UTC), datetime(2026, 1, 1, 0, 1, tzinfo=UTC)),
        ]
        col_names = ["client_order_id", "strategy_id", "symbol", "side", "qty", "order_type", "limit_price", "stop_price", "time_in_force", "status", "filled_qty", "filled_avg_price", "created_at", "filled_at"]
        cursor = _make_cursor_mock(rows, col_names)
        conn = _make_conn_mock(cursor)
        ctx = create_mock_context()
        ctx.db.transaction.return_value = conn

        _content, row_count = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="orders",
            strategy_ids=["alpha"],
            filter_params={"symbol": {"filterType": "text", "type": "equals", "filter": "AAPL"}},
            visible_columns=None,
            sort_model=None,
        )

        # The SQL executed should contain the filter clause.
        # psycopg.sql.Composed objects are passed to execute(); convert
        # to string representation for assertion.
        executed_sql = str(cursor.execute.call_args[0][0])
        assert "symbol = %s" in executed_sql
        # The bind parameters should include the filter value
        executed_params = cursor.execute.call_args[0][1]
        assert "AAPL" in executed_params

    async def test_sort_aliases_resolved_before_order_by(self) -> None:
        """Verify that frontend aliases in sort_model are resolved before ORDER BY."""
        rows = [
            ("ord-1", "alpha", "AAPL", "buy", 10, "market", None, None, "day", "filled", Decimal("10"), Decimal("150"), datetime(2026, 1, 1, tzinfo=UTC), datetime(2026, 1, 1, 0, 1, tzinfo=UTC)),
        ]
        col_names = ["client_order_id", "strategy_id", "symbol", "side", "qty", "order_type", "limit_price", "stop_price", "time_in_force", "status", "filled_qty", "filled_avg_price", "created_at", "filled_at"]
        cursor = _make_cursor_mock(rows, col_names)
        conn = _make_conn_mock(cursor)
        ctx = create_mock_context()
        ctx.db.transaction.return_value = conn

        _content, _row_count = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="orders",
            strategy_ids=["alpha"],
            filter_params=None,
            visible_columns=None,
            sort_model=[{"colId": "type", "sort": "desc"}],
        )

        # "type" should be resolved to "order_type" in the ORDER BY
        executed_sql = str(cursor.execute.call_args[0][0])
        assert "order_type DESC" in executed_sql

    async def test_audit_row_count_matches_data(self) -> None:
        """Verify row_count reflects actual exported rows, not a placeholder."""
        rows = [
            (1, datetime(2026, 1, 1, tzinfo=UTC), "admin", "login", "{}", None),
            (2, datetime(2026, 1, 2, tzinfo=UTC), "admin", "export", "{}", "test"),
            (3, datetime(2026, 1, 3, tzinfo=UTC), "admin", "logout", "{}", None),
        ]
        col_names = ["id", "timestamp", "user_id", "action", "details", "reason"]
        ctx = _make_ctx_with_rows(rows, col_names)

        _, row_count = await export_module._generate_excel_content(
            ctx=ctx,
            grid_name="audit",
            strategy_ids=["alpha"],
            filter_params=None,
            visible_columns=None,
            sort_model=None,
        )

        assert row_count == 3
