"""Tests for Excel export with real grid data (issue #151).

Validates that _generate_excel_content returns real data (not placeholders)
for all supported grids, applies filters/sort, respects column order,
and that the audit grid is disabled for safety.

Also includes route-level tests for the download_excel_export endpoint.
"""

from __future__ import annotations

import asyncio
import io
import json
from datetime import UTC, datetime
from decimal import Decimal
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import UUID, uuid4

import pytest

from apps.execution_gateway.routes.export import (
    ExportAuditCreateRequest,
    _apply_filters,
    _apply_sort,
    _coerce_cell_value,
    _fetch_audit_data,
    _fetch_fills_data,
    _fetch_orders_data,
    _fetch_positions_data,
    _fetch_tca_data,
    _generate_excel_content,
    _match_compound_filter,
    _match_filter,
    download_excel_export,
    router,
)
from libs.platform.security import sanitize_for_export


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


class _FakePosition:
    """Lightweight fake Position for testing."""

    def __init__(self, **kwargs: Any) -> None:
        defaults = {
            "symbol": "AAPL",
            "qty": Decimal("10"),
            "avg_entry_price": Decimal("150.25"),
            "current_price": Decimal("155.00"),
            "unrealized_pl": Decimal("47.50"),
            "realized_pl": Decimal("0"),
            "updated_at": datetime(2026, 3, 28, 16, 0, tzinfo=UTC),
            "last_trade_at": datetime(2026, 3, 28, 15, 30, tzinfo=UTC),
        }
        defaults.update(kwargs)
        for k, v in defaults.items():
            setattr(self, k, v)


def _make_position(**overrides: Any) -> _FakePosition:
    return _FakePosition(**overrides)


def _make_ctx(
    *,
    positions: list | None = None,
    fills: list[dict] | None = None,
    tca_trades: list[dict] | None = None,
    order_rows: list[tuple] | None = None,
) -> MagicMock:
    """Build a mock AppContext with db methods."""
    ctx = MagicMock()
    ctx.db.get_positions_for_strategies.return_value = positions or []
    ctx.db.get_recent_fills.return_value = fills or []
    ctx.db.get_trades_for_tca.return_value = tca_trades or []

    cursor = MagicMock()
    cursor.fetchall.return_value = order_rows or []
    cursor.__enter__ = MagicMock(return_value=cursor)
    cursor.__exit__ = MagicMock(return_value=False)

    conn = MagicMock()
    conn.cursor.return_value = cursor
    conn.__enter__ = MagicMock(return_value=conn)
    conn.__exit__ = MagicMock(return_value=False)

    ctx.db.transaction.return_value = conn
    return ctx


def _load_workbook(excel_bytes: bytes):
    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(excel_bytes))


def _run(coro):
    """Run an async coroutine synchronously."""
    return asyncio.get_event_loop().run_until_complete(coro)


# ---------------------------------------------------------------------------
# _coerce_cell_value tests
# ---------------------------------------------------------------------------


class TestCoerceCellValue:
    def test_none(self) -> None:
        assert _coerce_cell_value(None) is None

    def test_decimal_to_float(self) -> None:
        assert _coerce_cell_value(Decimal("3.14")) == pytest.approx(3.14)

    def test_datetime_strips_tzinfo(self) -> None:
        dt = datetime(2026, 1, 1, tzinfo=UTC)
        result = _coerce_cell_value(dt)
        assert isinstance(result, datetime)
        assert result.tzinfo is None

    def test_int_passthrough(self) -> None:
        assert _coerce_cell_value(42) == 42

    def test_string_sanitized(self) -> None:
        assert _coerce_cell_value("=SUM(A1)") == "'=SUM(A1)"

    def test_dict_json_sanitized(self) -> None:
        result = _coerce_cell_value({"key": "value"})
        assert '"key"' in result
        assert '"value"' in result


# ---------------------------------------------------------------------------
# _match_filter tests
# ---------------------------------------------------------------------------


class TestMatchFilter:
    def test_text_contains(self) -> None:
        f = {"filterType": "text", "type": "contains", "filter": "AA"}
        assert _match_filter("AAPL", f) is True
        assert _match_filter("MSFT", f) is False

    def test_text_equals(self) -> None:
        f = {"filterType": "text", "type": "equals", "filter": "AAPL"}
        assert _match_filter("AAPL", f) is True
        assert _match_filter("aapl", f) is True  # case-insensitive
        assert _match_filter("AAPLX", f) is False

    def test_text_starts_with(self) -> None:
        f = {"filterType": "text", "type": "startsWith", "filter": "MS"}
        assert _match_filter("MSFT", f) is True
        assert _match_filter("AAPL", f) is False

    def test_text_not_contains(self) -> None:
        f = {"filterType": "text", "type": "notContains", "filter": "X"}
        assert _match_filter("AAPL", f) is True
        assert _match_filter("XOM", f) is False

    def test_number_greater_than(self) -> None:
        f = {"filterType": "number", "type": "greaterThan", "filter": 50}
        assert _match_filter(100, f) is True
        assert _match_filter(Decimal("100"), f) is True
        assert _match_filter(30, f) is False

    def test_number_in_range(self) -> None:
        f = {"filterType": "number", "type": "inRange", "filter": 10, "filterTo": 50}
        assert _match_filter(25, f) is True
        assert _match_filter(5, f) is False
        assert _match_filter(60, f) is False

    def test_number_equals(self) -> None:
        f = {"filterType": "number", "type": "equals", "filter": 42}
        assert _match_filter(42, f) is True
        assert _match_filter(43, f) is False

    def test_none_value_matches_blank(self) -> None:
        f = {"filterType": "text", "type": "blank"}
        assert _match_filter(None, f) is True

    def test_none_value_rejects_non_blank(self) -> None:
        f = {"filterType": "text", "type": "contains", "filter": "X"}
        assert _match_filter(None, f) is False

    def test_date_equals(self) -> None:
        f = {"filterType": "date", "type": "equals", "dateFrom": "2026-03-28 00:00:00"}
        assert _match_filter(datetime(2026, 3, 28, 14, 30, tzinfo=UTC), f) is True
        assert _match_filter(datetime(2026, 3, 27, 23, 59, tzinfo=UTC), f) is False

    def test_date_not_equal(self) -> None:
        f = {"filterType": "date", "type": "notEqual", "dateFrom": "2026-03-28 00:00:00"}
        assert _match_filter(datetime(2026, 3, 28, 10, 0, tzinfo=UTC), f) is False
        assert _match_filter(datetime(2026, 3, 27, 10, 0, tzinfo=UTC), f) is True

    def test_date_greater_than(self) -> None:
        f = {"filterType": "date", "type": "greaterThan", "dateFrom": "2026-03-28 00:00:00"}
        assert _match_filter(datetime(2026, 3, 29, tzinfo=UTC), f) is True
        assert _match_filter(datetime(2026, 3, 27, tzinfo=UTC), f) is False

    def test_date_greater_than_same_day(self) -> None:
        """Same-day values must NOT pass greaterThan — AG Grid uses day granularity."""
        f = {"filterType": "date", "type": "greaterThan", "dateFrom": "2026-03-28 00:00:00"}
        # Same day at various times — all should be excluded
        assert _match_filter(datetime(2026, 3, 28, 0, 0, tzinfo=UTC), f) is False
        assert _match_filter(datetime(2026, 3, 28, 10, 30, tzinfo=UTC), f) is False
        assert _match_filter(datetime(2026, 3, 28, 23, 59, 59, tzinfo=UTC), f) is False
        # Next day — should pass
        assert _match_filter(datetime(2026, 3, 29, 0, 0, tzinfo=UTC), f) is True

    def test_date_less_than(self) -> None:
        f = {"filterType": "date", "type": "lessThan", "dateFrom": "2026-03-28 00:00:00"}
        assert _match_filter(datetime(2026, 3, 27, tzinfo=UTC), f) is True
        assert _match_filter(datetime(2026, 3, 29, tzinfo=UTC), f) is False

    def test_date_less_than_same_day(self) -> None:
        """Same-day values must NOT pass lessThan — AG Grid uses day granularity."""
        f = {"filterType": "date", "type": "lessThan", "dateFrom": "2026-03-28 00:00:00"}
        # Same day at various times — all should be excluded
        assert _match_filter(datetime(2026, 3, 28, 0, 0, tzinfo=UTC), f) is False
        assert _match_filter(datetime(2026, 3, 28, 10, 30, tzinfo=UTC), f) is False
        assert _match_filter(datetime(2026, 3, 28, 23, 59, 59, tzinfo=UTC), f) is False
        # Previous day — should pass
        assert _match_filter(datetime(2026, 3, 27, 23, 59, 59, tzinfo=UTC), f) is True

    def test_date_in_range(self) -> None:
        f = {
            "filterType": "date",
            "type": "inRange",
            "dateFrom": "2026-03-25 00:00:00",
            "dateTo": "2026-03-28 23:59:59",
        }
        assert _match_filter(datetime(2026, 3, 26, tzinfo=UTC), f) is True
        assert _match_filter(datetime(2026, 3, 24, tzinfo=UTC), f) is False
        assert _match_filter(datetime(2026, 3, 30, tzinfo=UTC), f) is False

    def test_date_in_range_boundary_days(self) -> None:
        """Boundary days of inRange must be included regardless of time-of-day."""
        f = {
            "filterType": "date",
            "type": "inRange",
            "dateFrom": "2026-03-25 00:00:00",
            "dateTo": "2026-03-28 00:00:00",
        }
        # Start boundary day at any time
        assert _match_filter(datetime(2026, 3, 25, 0, 0, tzinfo=UTC), f) is True
        assert _match_filter(datetime(2026, 3, 25, 23, 59, tzinfo=UTC), f) is True
        # End boundary day at any time
        assert _match_filter(datetime(2026, 3, 28, 0, 0, tzinfo=UTC), f) is True
        assert _match_filter(datetime(2026, 3, 28, 23, 59, tzinfo=UTC), f) is True
        # Outside range
        assert _match_filter(datetime(2026, 3, 24, 23, 59, tzinfo=UTC), f) is False
        assert _match_filter(datetime(2026, 3, 29, 0, 0, tzinfo=UTC), f) is False

    def test_date_blank_and_not_blank(self) -> None:
        f_blank = {"filterType": "date", "type": "blank"}
        assert _match_filter(None, f_blank) is True
        assert _match_filter(datetime(2026, 3, 28, tzinfo=UTC), f_blank) is False

        f_not_blank = {"filterType": "date", "type": "notBlank"}
        assert _match_filter(datetime(2026, 3, 28, tzinfo=UTC), f_not_blank) is True
        assert _match_filter(None, f_not_blank) is False

    def test_date_with_date_object(self) -> None:
        """date (not datetime) cell values must also match."""
        from datetime import date as date_type
        f = {"filterType": "date", "type": "equals", "dateFrom": "2026-03-28 00:00:00"}
        assert _match_filter(date_type(2026, 3, 28), f) is True
        assert _match_filter(date_type(2026, 3, 27), f) is False

    def test_date_with_iso_string_value(self) -> None:
        """String cell values that are ISO dates should be parsed and filtered."""
        f = {"filterType": "date", "type": "greaterThan", "dateFrom": "2026-03-28 00:00:00"}
        assert _match_filter("2026-03-29T10:00:00", f) is True
        assert _match_filter("2026-03-27T10:00:00", f) is False
        # Same-day ISO string must NOT pass greaterThan
        assert _match_filter("2026-03-28T15:00:00", f) is False

    def test_date_missing_date_from_keeps_row(self) -> None:
        """If dateFrom is missing, filter can't apply — keep the row."""
        f = {"filterType": "date", "type": "equals"}
        assert _match_filter(datetime(2026, 3, 28, tzinfo=UTC), f) is True

    def test_unknown_filter_type_keeps_row(self) -> None:
        f = {"filterType": "custom", "type": "magic", "filter": "x"}
        assert _match_filter("anything", f) is True


# ---------------------------------------------------------------------------
# _match_compound_filter tests
# ---------------------------------------------------------------------------


class TestMatchCompoundFilter:
    def test_and_compound_both_true(self) -> None:
        f = {
            "filterType": "number",
            "operator": "AND",
            "condition1": {"filterType": "number", "type": "greaterThan", "filter": 10},
            "condition2": {"filterType": "number", "type": "lessThan", "filter": 50},
        }
        assert _match_compound_filter(25, f) is True

    def test_and_compound_one_false(self) -> None:
        f = {
            "filterType": "number",
            "operator": "AND",
            "condition1": {"filterType": "number", "type": "greaterThan", "filter": 10},
            "condition2": {"filterType": "number", "type": "lessThan", "filter": 50},
        }
        assert _match_compound_filter(5, f) is False
        assert _match_compound_filter(60, f) is False

    def test_or_compound(self) -> None:
        f = {
            "filterType": "text",
            "operator": "OR",
            "condition1": {"filterType": "text", "type": "equals", "filter": "AAPL"},
            "condition2": {"filterType": "text", "type": "equals", "filter": "MSFT"},
        }
        assert _match_compound_filter("AAPL", f) is True
        assert _match_compound_filter("MSFT", f) is True
        assert _match_compound_filter("GOOG", f) is False

    def test_simple_filter_passthrough(self) -> None:
        """Non-compound filters delegate to _match_filter."""
        f = {"filterType": "text", "type": "contains", "filter": "AA"}
        assert _match_compound_filter("AAPL", f) is True
        assert _match_compound_filter("MSFT", f) is False


# ---------------------------------------------------------------------------
# _apply_filters tests
# ---------------------------------------------------------------------------


class TestApplyFilters:
    def test_text_filter_reduces_rows(self) -> None:
        columns = ["symbol", "qty"]
        rows = [["AAPL", 10], ["MSFT", 20], ["AMZN", 30]]
        filters = {"symbol": {"filterType": "text", "type": "contains", "filter": "A"}}
        result = _apply_filters(columns, rows, filters)
        assert len(result) == 2
        assert result[0][0] == "AAPL"
        assert result[1][0] == "AMZN"

    def test_number_filter(self) -> None:
        columns = ["symbol", "qty"]
        rows = [["AAPL", 10], ["MSFT", 20], ["AMZN", 30]]
        filters = {"qty": {"filterType": "number", "type": "greaterThan", "filter": 15}}
        result = _apply_filters(columns, rows, filters)
        assert len(result) == 2
        symbols = [r[0] for r in result]
        assert "AAPL" not in symbols

    def test_multiple_filters_combined(self) -> None:
        columns = ["symbol", "qty"]
        rows = [["AAPL", 10], ["AMZN", 30], ["AMAT", 5]]
        filters = {
            "symbol": {"filterType": "text", "type": "startsWith", "filter": "A"},
            "qty": {"filterType": "number", "type": "greaterThan", "filter": 8},
        }
        result = _apply_filters(columns, rows, filters)
        # All start with A, but only AAPL(10) and AMZN(30) have qty > 8
        assert len(result) == 2

    def test_unknown_column_ignored(self) -> None:
        columns = ["symbol"]
        rows = [["AAPL"], ["MSFT"]]
        filters = {"nonexistent": {"filterType": "text", "type": "equals", "filter": "X"}}
        result = _apply_filters(columns, rows, filters)
        assert len(result) == 2  # No filtering applied

    def test_empty_filter_params(self) -> None:
        columns = ["symbol"]
        rows = [["AAPL"]]
        result = _apply_filters(columns, rows, {})
        assert len(result) == 1

    def test_compound_and_filter(self) -> None:
        """AG Grid compound filter with AND operator must filter correctly."""
        columns = ["symbol", "qty"]
        rows = [["AAPL", 5], ["MSFT", 20], ["AMZN", 30], ["GOOG", 50]]
        filters = {
            "qty": {
                "filterType": "number",
                "operator": "AND",
                "condition1": {"filterType": "number", "type": "greaterThan", "filter": 10},
                "condition2": {"filterType": "number", "type": "lessThan", "filter": 40},
            }
        }
        result = _apply_filters(columns, rows, filters)
        # Only MSFT(20) and AMZN(30) are in range (10, 40)
        assert len(result) == 2
        assert result[0][0] == "MSFT"
        assert result[1][0] == "AMZN"

    def test_compound_or_filter(self) -> None:
        """AG Grid compound filter with OR operator."""
        columns = ["symbol", "qty"]
        rows = [["AAPL", 5], ["MSFT", 20], ["AMZN", 30]]
        filters = {
            "symbol": {
                "filterType": "text",
                "operator": "OR",
                "condition1": {"filterType": "text", "type": "equals", "filter": "AAPL"},
                "condition2": {"filterType": "text", "type": "equals", "filter": "AMZN"},
            }
        }
        result = _apply_filters(columns, rows, filters)
        assert len(result) == 2
        symbols = [r[0] for r in result]
        assert "AAPL" in symbols
        assert "AMZN" in symbols
        assert "MSFT" not in symbols


# ---------------------------------------------------------------------------
# _apply_sort tests
# ---------------------------------------------------------------------------


class TestApplySort:
    def test_sort_ascending(self) -> None:
        columns = ["symbol", "qty"]
        rows = [["MSFT", 20], ["AAPL", 10], ["AMZN", 30]]
        sort_model = [{"colId": "symbol", "sort": "asc"}]
        result = _apply_sort(columns, rows, sort_model)
        assert [r[0] for r in result] == ["AAPL", "AMZN", "MSFT"]

    def test_sort_descending(self) -> None:
        columns = ["symbol", "qty"]
        rows = [["MSFT", 20], ["AAPL", 10], ["AMZN", 30]]
        sort_model = [{"colId": "qty", "sort": "desc"}]
        result = _apply_sort(columns, rows, sort_model)
        assert [r[1] for r in result] == [30, 20, 10]

    def test_multi_sort(self) -> None:
        columns = ["side", "qty"]
        rows = [["buy", 20], ["sell", 10], ["buy", 5], ["sell", 30]]
        sort_model = [
            {"colId": "side", "sort": "asc"},
            {"colId": "qty", "sort": "desc"},
        ]
        result = _apply_sort(columns, rows, sort_model)
        # Primary: side asc (buy, buy, sell, sell)
        # Secondary: qty desc within each group
        assert result[0] == ["buy", 20]
        assert result[1] == ["buy", 5]
        assert result[2] == ["sell", 30]
        assert result[3] == ["sell", 10]

    def test_sort_nulls_last(self) -> None:
        columns = ["symbol"]
        rows = [["AAPL"], [None], ["MSFT"]]
        sort_model = [{"colId": "symbol", "sort": "asc"}]
        result = _apply_sort(columns, rows, sort_model)
        assert result[0] == ["AAPL"]
        assert result[1] == ["MSFT"]
        assert result[2] == [None]

    def test_sort_nulls_last_descending(self) -> None:
        """Nulls must sort last even with descending sort order."""
        columns = ["symbol"]
        rows = [["AAPL"], [None], ["MSFT"]]
        sort_model = [{"colId": "symbol", "sort": "desc"}]
        result = _apply_sort(columns, rows, sort_model)
        # Descending: MSFT, AAPL, then None last
        assert result[0] == ["MSFT"]
        assert result[1] == ["AAPL"]
        assert result[2] == [None]

    def test_unknown_column_ignored(self) -> None:
        columns = ["symbol"]
        rows = [["MSFT"], ["AAPL"]]
        sort_model = [{"colId": "nonexistent", "sort": "asc"}]
        result = _apply_sort(columns, rows, sort_model)
        # Order unchanged
        assert result[0] == ["MSFT"]

    def test_sort_index_overrides_array_order(self) -> None:
        """sortIndex must determine multi-sort precedence, not array position.

        Regression test: AG Grid serializes sortIndex in getColumnState().
        If the array order differs from sortIndex, the export must follow
        sortIndex (lower = higher priority).
        """
        columns = ["side", "qty"]
        rows = [["buy", 20], ["sell", 10], ["buy", 5], ["sell", 30]]

        # Array order: qty first, side second — but sortIndex says
        # side is primary (0) and qty is secondary (1).
        sort_model = [
            {"colId": "qty", "sort": "desc", "sortIndex": 1},
            {"colId": "side", "sort": "asc", "sortIndex": 0},
        ]
        result = _apply_sort(columns, rows, sort_model)
        # Primary: side asc (buy, buy, sell, sell)
        # Secondary: qty desc within each group
        assert result[0] == ["buy", 20]
        assert result[1] == ["buy", 5]
        assert result[2] == ["sell", 30]
        assert result[3] == ["sell", 10]

    def test_sort_index_missing_falls_back_to_array_order(self) -> None:
        """When sortIndex is absent, array position is used as fallback."""
        columns = ["side", "qty"]
        rows = [["buy", 20], ["sell", 10], ["buy", 5], ["sell", 30]]
        sort_model = [
            {"colId": "side", "sort": "asc"},
            {"colId": "qty", "sort": "desc"},
        ]
        result = _apply_sort(columns, rows, sort_model)
        # Same as test_multi_sort: side primary, qty secondary
        assert result[0] == ["buy", 20]
        assert result[1] == ["buy", 5]
        assert result[2] == ["sell", 30]
        assert result[3] == ["sell", 10]


# ---------------------------------------------------------------------------
# Per-grid fetcher tests
# ---------------------------------------------------------------------------


class TestFetchPositionsData:
    def test_returns_real_columns_and_rows(self) -> None:
        pos = _make_position()
        ctx = _make_ctx(positions=[pos])
        columns, rows = _fetch_positions_data(ctx, ["strat1"], None)

        assert "symbol" in columns
        assert "qty" in columns
        assert len(rows) == 1
        assert rows[0][columns.index("symbol")] == "AAPL"
        assert rows[0][columns.index("qty")] == Decimal("10")

    def test_empty_positions(self) -> None:
        ctx = _make_ctx(positions=[])
        columns, rows = _fetch_positions_data(ctx, ["strat1"], None)
        assert len(rows) == 0
        assert len(columns) > 0


class TestFetchOrdersData:
    def test_returns_real_rows(self) -> None:
        order_row = (
            "ord-123", "strat1", "MSFT", "buy", 100,
            "limit", "filled", Decimal("100"), Decimal("310.50"),
            datetime(2026, 3, 28, tzinfo=UTC),
            datetime(2026, 3, 28, tzinfo=UTC),
            datetime(2026, 3, 28, tzinfo=UTC),
        )
        ctx = _make_ctx(order_rows=[order_row])
        columns, rows = _fetch_orders_data(ctx, ["strat1"], None)

        assert "client_order_id" in columns
        assert len(rows) == 1
        assert rows[0][0] == "ord-123"


class TestFetchFillsData:
    def test_returns_real_fills(self) -> None:
        fill = {
            "client_order_id": "ord-456",
            "symbol": "GOOG",
            "side": "sell",
            "status": "filled",
            "qty": Decimal("5"),
            "price": Decimal("170.00"),
            "realized_pl": Decimal("25.00"),
            "timestamp": datetime(2026, 3, 28, tzinfo=UTC),
        }
        ctx = _make_ctx(fills=[fill])
        columns, rows = _fetch_fills_data(ctx, ["strat1"], None)

        assert len(rows) == 1
        assert rows[0][columns.index("symbol")] == "GOOG"


class TestFetchAuditData:
    def test_raises_not_implemented(self) -> None:
        """Audit export must be disabled — no strategy_id column for scoping."""
        ctx = _make_ctx()
        with pytest.raises(NotImplementedError, match="strategy"):
            _fetch_audit_data(ctx, ["strat1"], None)


class TestFetchTcaData:
    def test_raises_not_implemented(self) -> None:
        """TCA export must be disabled — UI uses computed columns the server can't reproduce."""
        ctx = _make_ctx()
        with pytest.raises(NotImplementedError, match="computed columns"):
            _fetch_tca_data(ctx, ["strat1"], None)


# ---------------------------------------------------------------------------
# Integration: _generate_excel_content end-to-end
# ---------------------------------------------------------------------------


class TestGenerateExcelContent:
    """Verify that _generate_excel_content produces real data workbooks."""

    def test_positions_grid_real_data(self) -> None:
        pos = _make_position(symbol="NVDA", qty=Decimal("50"))
        ctx = _make_ctx(positions=[pos])

        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=None,
                sort_model=None,
            )
        )

        assert row_count == 1
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        assert ws.title == "Positions"

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "symbol" in headers
        symbol_col = headers.index("symbol") + 1
        assert ws.cell(row=2, column=symbol_col).value == "NVDA"

    def test_orders_grid_real_data(self) -> None:
        order_row = (
            "ord-abc", "strat1", "AAPL", "buy", 10,
            "market", "filled", Decimal("10"), Decimal("150.00"),
            datetime(2026, 3, 28, tzinfo=UTC), None, None,
        )
        ctx = _make_ctx(order_rows=[order_row])

        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="orders",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=None,
                sort_model=None,
            )
        )

        assert row_count == 1
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "ord-abc"

    def test_fills_grid_real_data(self) -> None:
        fill = {
            "client_order_id": "ord-fill",
            "symbol": "TSLA",
            "side": "sell",
            "status": "filled",
            "qty": Decimal("3"),
            "price": Decimal("250.00"),
            "realized_pl": Decimal("15.00"),
            "timestamp": datetime(2026, 3, 28, tzinfo=UTC),
        }
        ctx = _make_ctx(fills=[fill])

        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="fills",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=None,
                sort_model=None,
            )
        )

        assert row_count == 1
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        symbol_col = headers.index("symbol") + 1
        assert ws.cell(row=2, column=symbol_col).value == "TSLA"

    def test_visible_columns_preserves_client_order(self) -> None:
        """Visible columns must appear in the order the client sent them."""
        pos = _make_position(symbol="META", qty=Decimal("100"))
        ctx = _make_ctx(positions=[pos])

        # Client sends columns in reversed order: qty first, then symbol
        excel_bytes, _ = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=["qty", "symbol"],
                sort_model=None,
            )
        )

        wb = _load_workbook(excel_bytes)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers == ["qty", "symbol"]
        # Verify data follows same order
        assert ws.cell(row=2, column=1).value == pytest.approx(100.0)  # qty (Decimal→float)
        assert ws.cell(row=2, column=2).value == "META"  # symbol

    def test_visible_columns_subset(self) -> None:
        pos = _make_position(symbol="META")
        ctx = _make_ctx(positions=[pos])

        excel_bytes, _ = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=["symbol", "qty"],
                sort_model=None,
            )
        )

        wb = _load_workbook(excel_bytes)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers == ["symbol", "qty"]
        assert ws.max_column == 2

    def test_filter_params_applied(self) -> None:
        """Filter params must reduce exported rows."""
        positions = [
            _make_position(symbol="AAPL", qty=Decimal("10")),
            _make_position(symbol="MSFT", qty=Decimal("20")),
            _make_position(symbol="AMZN", qty=Decimal("30")),
        ]
        ctx = _make_ctx(positions=positions)

        filter_params = {
            "symbol": {"filterType": "text", "type": "contains", "filter": "A"},
        }

        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=filter_params,
                visible_columns=None,
                sort_model=None,
            )
        )

        # AAPL and AMZN contain "A", MSFT does not
        assert row_count == 2
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        sym_col = headers.index("symbol") + 1
        exported_symbols = [
            ws.cell(row=r, column=sym_col).value
            for r in range(2, 2 + row_count)
        ]
        assert "AAPL" in exported_symbols
        assert "AMZN" in exported_symbols
        assert "MSFT" not in exported_symbols

    def test_sort_model_applied(self) -> None:
        """Sort model must reorder exported rows."""
        positions = [
            _make_position(symbol="MSFT", qty=Decimal("20")),
            _make_position(symbol="AAPL", qty=Decimal("10")),
            _make_position(symbol="AMZN", qty=Decimal("30")),
        ]
        ctx = _make_ctx(positions=positions)

        sort_model = [{"colId": "symbol", "sort": "asc"}]

        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=None,
                sort_model=sort_model,
            )
        )

        assert row_count == 3
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        sym_col = headers.index("symbol") + 1
        exported_symbols = [
            ws.cell(row=r, column=sym_col).value
            for r in range(2, 2 + row_count)
        ]
        assert exported_symbols == ["AAPL", "AMZN", "MSFT"]

    def test_filter_and_sort_combined(self) -> None:
        """Filter + sort applied together."""
        positions = [
            _make_position(symbol="MSFT", qty=Decimal("20")),
            _make_position(symbol="AAPL", qty=Decimal("10")),
            _make_position(symbol="AMZN", qty=Decimal("30")),
            _make_position(symbol="GOOG", qty=Decimal("15")),
        ]
        ctx = _make_ctx(positions=positions)

        # Filter: qty > 12 → MSFT(20), AMZN(30), GOOG(15)
        # Sort: qty desc → AMZN(30), MSFT(20), GOOG(15)
        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params={"qty": {"filterType": "number", "type": "greaterThan", "filter": 12}},
                visible_columns=["symbol", "qty"],
                sort_model=[{"colId": "qty", "sort": "desc"}],
            )
        )

        assert row_count == 3
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers == ["symbol", "qty"]
        symbols = [ws.cell(row=r, column=1).value for r in range(2, 5)]
        assert symbols == ["AMZN", "MSFT", "GOOG"]

    def test_row_count_reflects_filtered_data(self) -> None:
        """Row count must reflect post-filter count, not pre-filter."""
        positions = [_make_position(symbol=f"SYM{i}") for i in range(5)]
        ctx = _make_ctx(positions=positions)

        # Filter that matches none
        _, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params={"symbol": {"filterType": "text", "type": "equals", "filter": "NOMATCH"}},
                visible_columns=None,
                sort_model=None,
            )
        )
        assert row_count == 0

    def test_date_filter_on_orders_grid(self) -> None:
        """Date filter on created_at must correctly filter order rows."""
        order_rows = [
            (
                "ord-1", "strat1", "AAPL", "buy", 10,
                "market", "filled", Decimal("10"), Decimal("150.00"),
                datetime(2026, 3, 25, 10, 0, tzinfo=UTC), None, None,
            ),
            (
                "ord-2", "strat1", "MSFT", "buy", 20,
                "market", "filled", Decimal("20"), Decimal("300.00"),
                datetime(2026, 3, 28, 14, 0, tzinfo=UTC), None, None,
            ),
            (
                "ord-3", "strat1", "GOOG", "sell", 5,
                "market", "filled", Decimal("5"), Decimal("170.00"),
                datetime(2026, 3, 30, 8, 0, tzinfo=UTC), None, None,
            ),
        ]
        ctx = _make_ctx(order_rows=order_rows)

        # Filter: created_at > 2026-03-27 → should keep ord-2 and ord-3
        filter_params = {
            "created_at": {
                "filterType": "date",
                "type": "greaterThan",
                "dateFrom": "2026-03-27 00:00:00",
            }
        }

        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="orders",
                strategy_ids=["strat1"],
                filter_params=filter_params,
                visible_columns=None,
                sort_model=None,
            )
        )

        assert row_count == 2
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        oid_col = headers.index("client_order_id") + 1
        exported_ids = [ws.cell(row=r, column=oid_col).value for r in range(2, 2 + row_count)]
        assert "ord-1" not in exported_ids
        assert "ord-2" in exported_ids
        assert "ord-3" in exported_ids

    def test_date_in_range_filter_on_positions(self) -> None:
        """Date inRange filter on updated_at must correctly scope positions."""
        positions = [
            _make_position(symbol="OLD", updated_at=datetime(2026, 3, 20, tzinfo=UTC)),
            _make_position(symbol="MID", updated_at=datetime(2026, 3, 26, tzinfo=UTC)),
            _make_position(symbol="NEW", updated_at=datetime(2026, 3, 30, tzinfo=UTC)),
        ]
        ctx = _make_ctx(positions=positions)

        filter_params = {
            "updated_at": {
                "filterType": "date",
                "type": "inRange",
                "dateFrom": "2026-03-24 00:00:00",
                "dateTo": "2026-03-28 23:59:59",
            }
        }

        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=filter_params,
                visible_columns=["symbol", "updated_at"],
                sort_model=None,
            )
        )

        assert row_count == 1
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "MID"

    def test_audit_grid_raises_not_implemented(self) -> None:
        """Audit grid export must be disabled for safety (no strategy scoping)."""
        ctx = _make_ctx()
        with pytest.raises(NotImplementedError, match="strategy"):
            _run(
                _generate_excel_content(
                    ctx=ctx,
                    grid_name="audit",
                    strategy_ids=["strat1"],
                    filter_params=None,
                    visible_columns=None,
                    sort_model=None,
                )
            )

    def test_tca_grid_raises_not_implemented(self) -> None:
        """TCA grid export must be disabled — UI uses computed columns."""
        ctx = _make_ctx()
        with pytest.raises(NotImplementedError, match="computed columns"):
            _run(
                _generate_excel_content(
                    ctx=ctx,
                    grid_name="tca",
                    strategy_ids=["strat1"],
                    filter_params=None,
                    visible_columns=None,
                    sort_model=None,
                )
            )

    def test_unsupported_grid_raises(self) -> None:
        ctx = _make_ctx()
        with pytest.raises(NotImplementedError, match="not implemented"):
            _run(
                _generate_excel_content(
                    ctx=ctx,
                    grid_name="unknown_grid",
                    strategy_ids=["s1"],
                    filter_params=None,
                    visible_columns=None,
                    sort_model=None,
                )
            )

    def test_empty_data_returns_headers_only(self) -> None:
        ctx = _make_ctx(positions=[])

        excel_bytes, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=None,
                sort_model=None,
            )
        )

        assert row_count == 0
        wb = _load_workbook(excel_bytes)
        ws = wb.active
        assert ws.cell(row=1, column=1).value is not None
        assert ws.cell(row=2, column=1).value is None

    def test_formula_injection_sanitized(self) -> None:
        pos = _make_position(symbol="=CMD()")
        ctx = _make_ctx(positions=[pos])

        excel_bytes, _ = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=["symbol"],
                sort_model=None,
            )
        )

        wb = _load_workbook(excel_bytes)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "'=CMD()"

    def test_row_count_reflects_real_data(self) -> None:
        positions = [_make_position(symbol=f"SYM{i}") for i in range(5)]
        ctx = _make_ctx(positions=positions)

        _, row_count = _run(
            _generate_excel_content(
                ctx=ctx,
                grid_name="positions",
                strategy_ids=["strat1"],
                filter_params=None,
                visible_columns=None,
                sort_model=None,
            )
        )

        assert row_count == 5


# ---------------------------------------------------------------------------
# Route-level tests: download_excel_export
# ---------------------------------------------------------------------------


class TestDownloadExcelExportRoute:
    """Route-level tests for GET /api/v1/export/excel/{audit_id}."""

    def _make_audit_record(
        self,
        *,
        audit_id: UUID | None = None,
        user_id: str = "admin",
        grid_name: str = "positions",
        export_type: str = "excel",
        status: str = "pending",
        strategy_ids: list[str] | None = None,
        filter_params: dict | None = None,
        visible_columns: list[str] | None = None,
        sort_model: list[dict] | None = None,
    ) -> dict[str, Any]:
        return {
            "audit_id": audit_id or uuid4(),
            "user_id": user_id,
            "export_type": export_type,
            "grid_name": grid_name,
            "filter_params": filter_params,
            "visible_columns": visible_columns,
            "sort_model": sort_model,
            "strategy_ids": strategy_ids or ["strat1"],
            "export_scope": "visible",
            "estimated_row_count": None,
            "actual_row_count": None,
            "reported_by": None,
            "status": status,
            "error_message": None,
            "ip_address": "127.0.0.1",
            "session_id": "sess-1",
            "user_agent": "test",
            "created_at": datetime(2026, 3, 28, tzinfo=UTC),
            "completed_at": None,
        }

    @pytest.mark.asyncio
    async def test_successful_download_returns_excel(self) -> None:
        """Happy path: returns StreamingResponse with Excel content."""
        audit_id = uuid4()
        audit_record = self._make_audit_record(audit_id=audit_id)
        pos = _make_position(symbol="TEST")
        ctx = _make_ctx(positions=[pos])
        user = {"user_id": "admin", "user": MagicMock(), "session_id": "sess-1"}

        with (
            patch(
                "apps.execution_gateway.routes.export._get_export_audit",
                new_callable=AsyncMock,
                return_value=audit_record,
            ),
            patch(
                "apps.execution_gateway.routes.export._claim_export_audit",
                new_callable=AsyncMock,
                return_value=True,
            ),
            patch(
                "apps.execution_gateway.routes.export._complete_and_expire_export_audit",
                new_callable=AsyncMock,
            ) as mock_complete,
        ):
            response = await download_excel_export(
                audit_id=audit_id,
                ctx=ctx,
                user=user,
                _auth_context=MagicMock(),
            )

        assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # Verify audit was completed with correct row count
        mock_complete.assert_called_once()
        call_kwargs = mock_complete.call_args
        assert call_kwargs[1]["actual_row_count"] == 1

    @pytest.mark.asyncio
    async def test_audit_grid_returns_501(self) -> None:
        """Audit grid export must return 501 (disabled for safety)."""
        from fastapi import HTTPException

        audit_id = uuid4()
        audit_record = self._make_audit_record(
            audit_id=audit_id, grid_name="audit"
        )
        ctx = _make_ctx()
        user = {"user_id": "admin", "user": MagicMock(), "session_id": "sess-1"}

        with (
            patch(
                "apps.execution_gateway.routes.export._get_export_audit",
                new_callable=AsyncMock,
                return_value=audit_record,
            ),
            patch(
                "apps.execution_gateway.routes.export._claim_export_audit",
                new_callable=AsyncMock,
                return_value=True,
            ),
            patch(
                "apps.execution_gateway.routes.export._fail_export_audit",
                new_callable=AsyncMock,
            ),
        ):
            with pytest.raises(HTTPException) as exc_info:
                await download_excel_export(
                    audit_id=audit_id,
                    ctx=ctx,
                    user=user,
                    _auth_context=MagicMock(),
                )
            assert exc_info.value.status_code == 501
            assert "strategy" in str(exc_info.value.detail).lower()

    @pytest.mark.asyncio
    async def test_wrong_user_returns_403(self) -> None:
        from fastapi import HTTPException

        audit_id = uuid4()
        audit_record = self._make_audit_record(
            audit_id=audit_id, user_id="other_user"
        )
        ctx = _make_ctx()
        user = {"user_id": "admin", "user": MagicMock(), "session_id": "sess-1"}

        with patch(
            "apps.execution_gateway.routes.export._get_export_audit",
            new_callable=AsyncMock,
            return_value=audit_record,
        ):
            with pytest.raises(HTTPException) as exc_info:
                await download_excel_export(
                    audit_id=audit_id,
                    ctx=ctx,
                    user=user,
                    _auth_context=MagicMock(),
                )
            assert exc_info.value.status_code == 403

    @pytest.mark.asyncio
    async def test_already_used_returns_410(self) -> None:
        from fastapi import HTTPException

        audit_id = uuid4()
        audit_record = self._make_audit_record(
            audit_id=audit_id, status="expired"
        )
        ctx = _make_ctx()
        user = {"user_id": "admin", "user": MagicMock(), "session_id": "sess-1"}

        with patch(
            "apps.execution_gateway.routes.export._get_export_audit",
            new_callable=AsyncMock,
            return_value=audit_record,
        ):
            with pytest.raises(HTTPException) as exc_info:
                await download_excel_export(
                    audit_id=audit_id,
                    ctx=ctx,
                    user=user,
                    _auth_context=MagicMock(),
                )
            assert exc_info.value.status_code == 410

    @pytest.mark.asyncio
    async def test_filter_and_sort_respected_in_route(self) -> None:
        """Route must pass filter/sort from audit record to Excel generator."""
        audit_id = uuid4()
        positions = [
            _make_position(symbol="AAPL", qty=Decimal("10")),
            _make_position(symbol="MSFT", qty=Decimal("20")),
            _make_position(symbol="AMZN", qty=Decimal("30")),
        ]
        audit_record = self._make_audit_record(
            audit_id=audit_id,
            filter_params={"symbol": {"filterType": "text", "type": "contains", "filter": "A"}},
            sort_model=[{"colId": "qty", "sort": "desc"}],
            visible_columns=["qty", "symbol"],
        )
        ctx = _make_ctx(positions=positions)
        user = {"user_id": "admin", "user": MagicMock(), "session_id": "sess-1"}

        with (
            patch(
                "apps.execution_gateway.routes.export._get_export_audit",
                new_callable=AsyncMock,
                return_value=audit_record,
            ),
            patch(
                "apps.execution_gateway.routes.export._claim_export_audit",
                new_callable=AsyncMock,
                return_value=True,
            ),
            patch(
                "apps.execution_gateway.routes.export._complete_and_expire_export_audit",
                new_callable=AsyncMock,
            ) as mock_complete,
        ):
            response = await download_excel_export(
                audit_id=audit_id,
                ctx=ctx,
                user=user,
                _auth_context=MagicMock(),
            )

        # Read the response body
        body = b""
        async for chunk in response.body_iterator:
            body += chunk if isinstance(chunk, bytes) else chunk.encode()

        wb = _load_workbook(body)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Client column order: qty, symbol
        assert headers == ["qty", "symbol"]

        # Filter: AAPL and AMZN (contain "A"), sort: qty desc → AMZN(30), AAPL(10)
        assert mock_complete.call_args[1]["actual_row_count"] == 2
        assert ws.cell(row=2, column=2).value == "AMZN"  # symbol col = 2
        assert ws.cell(row=3, column=2).value == "AAPL"
